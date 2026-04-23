from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timedelta
import openpyxl
import io
import re

from app.database import get_db
from app.models import Schedule, Event, Lecturer, EventType

router = APIRouter()

DAYS_MAP = {
    "poniedziałek": 0, "wtorek": 1, "środa": 2,
    "czwartek": 3, "piątek": 4, "sobota": 5, "niedziela": 6
}

DAYS_PL = {v: k for k, v in DAYS_MAP.items()}


# ----- Schedules CRUD -----

class ScheduleCreate(BaseModel):
    name: str
    semester: Optional[str] = None


@router.post("/schedules")
def create_schedule(data: ScheduleCreate, db: Session = Depends(get_db)):
    schedule = Schedule(name=data.name, semester=data.semester)
    db.add(schedule)
    db.commit()
    db.refresh(schedule)
    return {"id": schedule.id, "name": schedule.name, "semester": schedule.semester}


@router.get("/schedules")
def list_schedules(db: Session = Depends(get_db)):
    schedules = db.query(Schedule).all()
    return [
        {"id": s.id, "name": s.name, "semester": s.semester, "events_count": len(s.events)}
        for s in schedules
    ]


@router.delete("/schedules/{schedule_id}")
def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Plan nie istnieje")
    db.delete(schedule)
    db.commit()
    return {"message": f"Usunięto plan {schedule.name}"}


# ----- XLSX upload -----

@router.post("/schedules/{schedule_id}/upload")
def upload_schedule(
    schedule_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Plan nie istnieje")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Tylko pliki .xlsx")

    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    has_date = "Data" in headers

    # Wywal stare zajecia z tego planu
    db.query(Event).filter(
        Event.schedule_id == schedule_id,
        Event.type == EventType.zajecia,
    ).delete(synchronize_session=False)

    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        if has_date:
            przedmiot, dzien, data, od, do_, sala, prowadzacy, uwagi = (list(row) + [None] * 8)[:8]
        else:
            przedmiot, dzien, od, do_, sala, prowadzacy, uwagi = (list(row) + [None] * 7)[:7]
            data = None

        date_obj = None
        if data:
            if isinstance(data, str):
                try:
                    date_obj = datetime.strptime(data, "%Y-%m-%d")
                except ValueError:
                    pass
            elif hasattr(data, "year"):
                date_obj = datetime(data.year, data.month, data.day)

        event = Event(
            type=EventType.zajecia,
            title=str(przedmiot) if przedmiot else "(bez tytułu)",
            day_of_week=str(dzien).lower().strip() if dzien else None,
            date=date_obj,
            time_start=str(od) if od else None,
            time_end=str(do_) if do_ else None,
            location=str(sala) if sala else None,
            lecturer=str(prowadzacy) if prowadzacy else None,
            notes=str(uwagi) if uwagi else None,
            schedule_id=schedule_id,
        )
        db.add(event)
        added += 1

    db.commit()
    return {"message": f"Dodano {added} zajęć do planu {schedule.name}"}


# ----- ICS import -----

SKROTY = {
    "Gk": "Grafika komputerowa",
    "Iwpp GJ": "Informatyka w procesach produkcyjnych",
    "Iwpp PG": "Informatyka w procesach produkcyjnych",
    "Prir": "Programowanie równoległe i rozproszone",
    "Prir - AK": "Programowanie równoległe i rozproszone - projekt AK",
    "Ps": "Programowanie systemowe",
    "Smiw": "Systemy mikroprocesorowe i wbudowane",
    "Smiw w": "Systemy mikroprocesorowe i wbudowane",
    "Taiib": "Tworzenie aplikacji internetowych i bazodanowych",
    "Taiib - P": "Tworzenie aplikacji internetowych i bazodanowych - projekt",
    "Wtp": "Współczesne techniki programowania",
    "WPP": "Wizualizacja procesów przemysłowych",
    "WPP - GK": "Wizualizacja procesów przemysłowych - projekt GK",
    "IO": "Inżynieria oprogramowania",
}


def detect_type(summary: str) -> EventType:
    s = summary.lower()
    if "egzamin" in s or "test" in s:
        return EventType.egzamin
    if "proj" in s:
        return EventType.projekt
    if "wyk" in s:
        return EventType.wyklad
    if "lab" in s:
        return EventType.laboratorium
    if "online" in s or "zdaln" in s:
        return EventType.online
    if "cwicz" in s or "ćwicz" in s:
        return EventType.cwiczenia
    return EventType.zajecia


@router.post("/schedules/{schedule_id}/import-ics")
async def import_ics(
    schedule_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Plan nie istnieje")

    content = (await file.read()).decode("utf-8")
    warsaw_offset = timedelta(hours=2)

    blocks = content.split("BEGIN:VEVENT")[1:]
    added = 0
    skipped = 0

    for block in blocks:
        summary_match = re.search(r"SUMMARY:(.*?)(?:\r\n|\n)", block)
        start_match = re.search(r"DTSTART:(\d{8}T\d{6}Z)", block)
        end_match = re.search(r"DTEND:(\d{8}T\d{6}Z)", block)

        if not (summary_match and start_match and end_match):
            continue

        summary = summary_match.group(1).strip()
        summary_lower = summary.lower()
        if (
            "wakacje" in summary_lower
            or "test zdalny" in summary_lower
            or "dni świąteczne" in summary_lower
            or "dni rektorskie" in summary_lower
            or "dzień rektorski" in summary_lower
        ):
            skipped += 1
            continue

        dt_start = datetime.strptime(start_match.group(1), "%Y%m%dT%H%M%SZ") + warsaw_offset
        dt_end = datetime.strptime(end_match.group(1), "%Y%m%dT%H%M%SZ") + warsaw_offset

        existing = db.query(Event).filter(
            Event.schedule_id == schedule_id,
            Event.date == dt_start.replace(hour=0, minute=0, second=0, microsecond=0),
            Event.time_start == dt_start.strftime("%H:%M"),
        ).first()
        if existing:
            skipped += 1
            continue

        parts = summary.split()
        type_idx = -1
        for i, p in enumerate(parts):
            if p in ["wyk", "lab", "proj", "sem", "w"]:
                type_idx = i
                break

        if type_idx >= 0:
            skrot = " ".join(parts[:type_idx])
            rest = parts[type_idx + 1:]
        else:
            skrot = summary
            rest = []

        prowadzacy = ""
        sala_parts = rest
        if rest and len(rest[0]) <= 5 and rest[0][0].isupper():
            prowadzacy = rest[0]
            sala_parts = rest[1:]

        sala = " ".join(sala_parts).strip()
        przedmiot = SKROTY.get(skrot, skrot)

        event = Event(
            type=detect_type(summary),
            title=przedmiot,
            date=dt_start.replace(hour=0, minute=0, second=0, microsecond=0),
            day_of_week=DAYS_PL[dt_start.weekday()],
            time_start=dt_start.strftime("%H:%M"),
            time_end=dt_end.strftime("%H:%M"),
            location=sala,
            lecturer=prowadzacy,
            schedule_id=schedule_id,
        )
        db.add(event)
        added += 1

    db.commit()
    return {"added": added, "skipped": skipped}


# ----- Events CRUD (edycja z weba) -----

class EventCreate(BaseModel):
    type: EventType = EventType.zajecia
    title: str
    day_of_week: Optional[str] = None
    date: Optional[str] = None  # "YYYY-MM-DD"
    time_start: Optional[str] = None
    time_end: Optional[str] = None
    location: Optional[str] = None
    lecturer: Optional[str] = None
    notes: Optional[str] = None


class EventUpdate(BaseModel):
    type: Optional[EventType] = None
    title: Optional[str] = None
    day_of_week: Optional[str] = None
    date: Optional[str] = None
    time_start: Optional[str] = None
    time_end: Optional[str] = None
    location: Optional[str] = None
    lecturer: Optional[str] = None
    notes: Optional[str] = None
    is_cancelled: Optional[bool] = None


def _parse_date(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Data musi być w formacie YYYY-MM-DD")


def _event_to_dict(e: Event) -> dict:
    return {
        "id": e.id,
        "type": e.type.value if e.type else None,
        "title": e.title,
        "day_of_week": e.day_of_week,
        "date": e.date.strftime("%Y-%m-%d") if e.date else None,
        "time_start": e.time_start,
        "time_end": e.time_end,
        "location": e.location,
        "lecturer": e.lecturer,
        "notes": e.notes,
        "is_cancelled": e.is_cancelled,
        "schedule_id": e.schedule_id,
    }


@router.get("/schedules/{schedule_id}/events")
def list_events(
    schedule_id: int,
    search: Optional[str] = None,
    day: Optional[str] = None,
    include_cancelled: bool = False,
    db: Session = Depends(get_db),
):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Plan nie istnieje")

    query = db.query(Event).filter(Event.schedule_id == schedule_id)
    if not include_cancelled:
        query = query.filter(Event.is_cancelled == False)  # noqa: E712
    if search:
        query = query.filter(Event.title.ilike(f"%{search}%"))
    if day:
        query = query.filter(Event.day_of_week == day.lower().strip())

    events = query.order_by(Event.date, Event.day_of_week, Event.time_start).all()
    return [_event_to_dict(e) for e in events]


@router.post("/schedules/{schedule_id}/events")
def create_event(
    schedule_id: int,
    data: EventCreate,
    db: Session = Depends(get_db),
):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Plan nie istnieje")

    event = Event(
        type=data.type,
        title=data.title,
        day_of_week=data.day_of_week.lower().strip() if data.day_of_week else None,
        date=_parse_date(data.date),
        time_start=data.time_start,
        time_end=data.time_end,
        location=data.location,
        lecturer=data.lecturer,
        notes=data.notes,
        schedule_id=schedule_id,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    return _event_to_dict(event)


@router.patch("/events/{event_id}")
def update_event(event_id: int, data: EventUpdate, db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Zajęcia nie istnieją")

    payload = data.model_dump(exclude_unset=True)
    if "date" in payload:
        payload["date"] = _parse_date(payload["date"])
    if "day_of_week" in payload and payload["day_of_week"]:
        payload["day_of_week"] = payload["day_of_week"].lower().strip()

    for key, value in payload.items():
        setattr(event, key, value)

    db.commit()
    db.refresh(event)
    return _event_to_dict(event)


@router.delete("/events/{event_id}")
def delete_event(event_id: int, db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Zajęcia nie istnieją")
    db.delete(event)
    db.commit()
    return {"message": "Usunięto"}


# ----- Lecturers -----

class LecturerCreate(BaseModel):
    abbreviation: str
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    room: Optional[str] = None
    phone: Optional[str] = None
    office_hours: Optional[str] = None


@router.post("/lecturers")
def create_lecturer(data: LecturerCreate, db: Session = Depends(get_db)):
    existing = db.query(Lecturer).filter(
        Lecturer.abbreviation == data.abbreviation
    ).first()
    if existing:
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(existing, key, value)
        db.commit()
        db.refresh(existing)
        return existing
    lecturer = Lecturer(**data.model_dump())
    db.add(lecturer)
    db.commit()
    db.refresh(lecturer)
    return lecturer


@router.get("/lecturers")
def list_lecturers(db: Session = Depends(get_db)):
    return db.query(Lecturer).all()


@router.get("/lecturers/{abbreviation}")
def get_lecturer(abbreviation: str, db: Session = Depends(get_db)):
    lecturer = db.query(Lecturer).filter(
        Lecturer.abbreviation == abbreviation
    ).first()
    if not lecturer:
        raise HTTPException(status_code=404, detail="Prowadzący nie znaleziony")
    return lecturer